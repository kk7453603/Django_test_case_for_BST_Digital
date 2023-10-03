import openpyxl
from django.http import HttpResponse
import datetime
from robots.models import Robot
from django.db.models import Count


def export_production(request):
    last_week = datetime.date.today() - datetime.timedelta(days=7)
    robot_models = Robot.objects.filter(created__gte=str(last_week), created__lte=str(datetime.date.today())).values(
        'model', 'version').annotate(count=Count('id'))
    # суть запроса на SQL: SELECT model,version,COUNT(*) FROM robots_robot GROUP BY serial,model

    print(robot_models)

    wb = openpyxl.Workbook()
    check_set = set()  # Костыль для неработающего DISTINCT
    for robot_model in robot_models:
        model = robot_model['model']
        if not model in check_set:
            check_set.add(model)
        else:
            continue
        wt = wb.create_sheet(title=model)
        wt['A1'] = 'Модель'
        wt['B1'] = 'Версия'
        wt['C1'] = 'Количество за неделю'

        robot_production = robot_models.filter(model=model)
        row = 2
        for rp in robot_production:
            wt.cell(row=row, column=1, value=rp['model'])
            wt.cell(row=row, column=2, value=rp['version'])
            wt.cell(row=row, column=3, value=rp['count'])
            row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=robot_production.xlsx'
    wb.save(response)
    return response
